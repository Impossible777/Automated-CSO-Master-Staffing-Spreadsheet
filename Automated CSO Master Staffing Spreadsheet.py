import openpyxl

# Load the source and destination workbooks
source_wb = openpyxl.load_workbook('Monday 4.2.xlsm')
dest_wb = openpyxl.load_workbook('Master Staffing Spreadsheet.xlsx')

# Select the source and destination worksheets
source_ws = source_wb['sheet1']
dest_ws = dest_wb['Monday 4.2']

# Specify the row and column to start copying to in the destination spreadsheet
dest_row = 121
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ED' or row[2] == 'BJC/CHS/CSO/BJC RN/RN-ICU':
            if row[8] == 'preassign':
                if row [2] == 'BJC/CHS/CSO/BJC RN/RN-ED':
                    dest_ws.cell(row=dest_row, column=dest_col+2).value = 'ED'
                elif row [2] == 'BJC/CHS/CSO/BJC RN/RN-ICU':
                     dest_ws.cell(row=dest_row, column=dest_col+2).value = 'ICU'   
                # Get the values from column D and E in the same row and copy them to the destination columns
                value_D = row[3]
                value_E = row[4]
                value_F = row[9]
                dest_ws.cell(row=dest_row, column=dest_col).value = value_D
                dest_ws.cell(row=dest_row, column=dest_col+1).value = value_E
                dest_ws.cell(row=dest_row, column=dest_col+9).value = value_F
                # Increment the destination row index
                dest_row += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row2 = 41
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ED'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row2, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row2, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row2, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row2, column=dest_col+2).value = 'ED'
            # Increment the destination row index
            dest_row2 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row3 = 47
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ICU'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row3, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row3, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row3, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row3, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row3 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row4 = 52
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row4, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row4, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row4, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row4, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row4 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row5 = 62
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/Support Staff/OBS'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row5, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row5, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row5, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row5, column=dest_col+2).value = 'Patient Observer'
            # Increment the destination row index
            dest_row5 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row6 = 68
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/Support Staff/PCT' and not row[8] or row[2] == 'BJC/CHS/CSO/Support Staff/SNT' and not row[8]:
            if row[2] == 'BJC/CHS/CSO/Support Staff/PCT':
                dest_ws.cell(row=dest_row6, column=dest_col+2).value = 'PCT'
            elif row[2] == 'BJC/CHS/CSO/Support Staff/SNT':
                dest_ws.cell(row=dest_row6, column=dest_col+2).value = 'SNT'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row6, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row6, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row6, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row6 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row7 = 79
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/Support Staff/PSA'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row7, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row7, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row7, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row7, column=dest_col+2).value = 'PSA'
            # Increment the destination row index
            dest_row7 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row8 = 86
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/IL-Agency/RN-ICU-IL'and not row[8] or row[2] == 'BJC/CHS/CSO/IL-Agency/RN-IL'and not row[8]:
            if row[2] == 'BJC/CHS/CSO/IL-Agency/RN-ICU-IL':
                dest_ws.cell(row=dest_row8, column=dest_col+2).value = 'ICU'
            elif row[2] == 'BJC/CHS/CSO/IL-Agency/RN-IL':
                dest_ws.cell(row=dest_row8, column=dest_col+2).value = 'Non-ICU'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row8, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row8, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row8, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row8 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row9 = 89
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-ICU-MO'and not row[8] or row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-MO'and not row[8]:
            if row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-ICU-MO':
                dest_ws.cell(row=dest_row9, column=dest_col+2).value = 'ICU'
            elif row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-MO':
                dest_ws.cell(row=dest_row9, column=dest_col+2).value = 'Non-ICU'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row9, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row9, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row9, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row9 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row10 = 95
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/R1-Agency/RN-ICU-MO'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row10, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row10, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row10, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row10, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row10 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row11 = 98
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/R2-Agency/RN-ICU-MO'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row11, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row11, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row11, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row11, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row11 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row12 = 107
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/R3-Agency/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row12, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row12, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row12, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row12, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row12 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row13 = 111
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[2] == 'BJC/CHS/CSO/R4-Agency/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row13, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row13, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row13, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row13, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row13 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row14 = 154
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '7:00 - 15:30' or  row[4] == '7:00 - 19:30' or row [4] == '5:45 - 18:15' or row[4] == '6:00 - 18:30' or row[4] == '6:45 - 19:15' or row[4] == '11:00 - 23:30' or row[4] == '7:00 - 11:30':
        if row[6] == 'BJC/CHS/CSO/EH/OBS' or row[6] == 'BJC/CHS/CSO/EH/PCT' or row[6] == 'BJC/CHS/CSO/EH/RN':
            if row[6] == 'BJC/CHS/CSO/EH/OBS':
                dest_ws.cell(row=dest_row14, column=dest_col+2).value = 'Patient Observer'
            elif row[6] == 'BJC/CHS/CSO/EH/PCT':
                dest_ws.cell(row=dest_row14, column=dest_col+2).value = 'PCT'
            elif row[6] == 'BJC/CHS/CSO/EH/RN':
                dest_ws.cell(row=dest_row14, column=dest_col+2).value = 'RN'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row14, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row14, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row14, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row14 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row15 = 276
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ED' or row[2] == 'BJC/CHS/CSO/BJC RN/RN-ICU' or row[2] == 'BJC/CHS/CSO/IL-Agency/RN-IL':
            if row[8] == 'preassign':
                if row [2] == 'BJC/CHS/CSO/BJC RN/RN-ED':
                    dest_ws.cell(row=dest_row15, column=dest_col+2).value = 'ED'
                elif row [2] == 'BJC/CHS/CSO/BJC RN/RN-ICU':
                     dest_ws.cell(row=dest_row15, column=dest_col+2).value = 'ICU'   
                # Get the values from column D and E in the same row and copy them to the destination columns
                value_D = row[3]
                value_E = row[4]
                value_F = row[9]
                dest_ws.cell(row=dest_row15, column=dest_col).value = value_D
                dest_ws.cell(row=dest_row15, column=dest_col+1).value = value_E
                dest_ws.cell(row=dest_row15, column=dest_col+9).value = value_F
                # Increment the destination row index
                dest_row15 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row16 = 202
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ED'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row16, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row16, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row16, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row16, column=dest_col+2).value = 'ED'
            # Increment the destination row index
            dest_row16 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row17 = 208
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN-ICU'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row17, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row17, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row17, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row17, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row17 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row18 = 214
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/BJC RN/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row18, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row18, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row18, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row18, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row18 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row19 = 225
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/Support Staff/OBS'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row19, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row19, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row19, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row19, column=dest_col+2).value = 'Patient Observer'
            # Increment the destination row index
            dest_row19 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row20 = 231
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/Support Staff/PCT' and not row[8] or row[2] == 'BJC/CHS/CSO/Support Staff/SNT' and not row[8]:
            if row[2] == 'BJC/CHS/CSO/Support Staff/PCT':
                dest_ws.cell(row=dest_row20, column=dest_col+2).value = 'PCT'
            elif row[2] == 'BJC/CHS/CSO/Support Staff/SNT':
                dest_ws.cell(row=dest_row20, column=dest_col+2).value = 'SNT'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row20, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row20, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row20, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row20 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row21 = 236
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/Support Staff/PSA'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row21, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row21, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row21, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row21, column=dest_col+2).value = 'PSA'
            # Increment the destination row index
            dest_row21 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row22 = 243
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/IL-Agency/RN-ICU-IL'and not row[8] or row[2] == 'BJC/CHS/CSO/IL-Agency/RN-IL'and not row[8]:
            if row[2] == 'BJC/CHS/CSO/IL-Agency/RN-ICU-IL':
                dest_ws.cell(row=dest_row22, column=dest_col+2).value = 'ICU'
            elif row[2] == 'BJC/CHS/CSO/IL-Agency/RN-IL':
                dest_ws.cell(row=dest_row22, column=dest_col+2).value = 'Non-ICU'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row22, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row22, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row22, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row22 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row23 = 246
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-ICU-MO'and not row[8] or row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-MO'and not row[8]:
            if row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-ICU-MO':
                dest_ws.cell(row=dest_row23, column=dest_col+2).value = 'ICU'
            elif row[2] == 'BJC/CHS/CSO/Metro-Agency/RN-MO':
                dest_ws.cell(row=dest_row23, column=dest_col+2).value = 'Non-ICU'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row23, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row23, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row23, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row23 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row24 = 252
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
   if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/R1-Agency/RN-ICU-MO'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row24, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row24, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row24, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row24, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row24 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row25 = 255
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/R2-Agency/RN-ICU-MO'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row25, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row25, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row25, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row25, column=dest_col+2).value = 'ICU'
            # Increment the destination row index
            dest_row25 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row26 = 264
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
  if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/R3-Agency/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row26, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row26, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row26, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row26, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row26 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row27 = 267
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[2] == 'BJC/CHS/CSO/R4-Agency/RN'and not row[8]:
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row27, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row27, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row27, column=dest_col+9).value = value_F
            dest_ws.cell(row=dest_row27, column=dest_col+2).value = 'Non-ICU'
            # Increment the destination row index
            dest_row27 += 1

# Specify the row and column to start copying to in the destination spreadsheet
dest_row28 = 307
dest_col = 2

# Loop through each row in the source spreadsheet
for i, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True)):
    # Check if the cell values in columns B and H are both "Apple"
    if  row[4] == '19:00 - 3:30' or  row[4] == '19:00 - 7:30' or row [4] == '17:45 - 6:15' or row[4] == '18:00 - 6:30' or row[4] == '18:45 - 7:15' or row[4] == '23:00 - 11:30' or row[4] == '19:00 - 23:30' or row[4] == '23:00 - 7:30' or row[4] == '21:00 - 6:30' or row[4] == '18:00 - 21:00':
        if row[6] == 'BJC/CHS/CSO/EH/OBS' or row[6] == 'BJC/CHS/CSO/EH/PCT' or row[6] == 'BJC/CHS/CSO/EH/RN':
            if row[6] == 'BJC/CHS/CSO/EH/OBS':
                dest_ws.cell(row=dest_row28, column=dest_col+2).value = 'Patient Observer'
            elif row[6] == 'BJC/CHS/CSO/EH/PCT':
                dest_ws.cell(row=dest_row28, column=dest_col+2).value = 'PCT'
            elif row[6] == 'BJC/CHS/CSO/EH/RN':
                dest_ws.cell(row=dest_row28, column=dest_col+2).value = 'RN'
            # Get the values from column D and E in the same row and copy them to the destination columns
            value_D = row[3]
            value_E = row[4]
            value_F = row[9]
            dest_ws.cell(row=dest_row28, column=dest_col).value = value_D
            dest_ws.cell(row=dest_row28, column=dest_col+1).value = value_E
            dest_ws.cell(row=dest_row28, column=dest_col+9).value = value_F
            # Increment the destination row index
            dest_row28 += 1

                
# Save the destination workbook
dest_wb.save('Master Staffing Spreadsheet.xlsx')
